import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import os
import excel2img

db_folder = '../lib/db/'
result_folder = '../results/'

def get_room_choice(code: str):
    if os.path.exists(f'{db_folder}room.json'):
        with open(f'{db_folder}room.json', 'r') as file:
            room_data = json.load(file)
        room_choice = room_data[code]
        return room_choice
    else:
        return None
    
def append_room_choice(room_choice: str, code: str):
    if not os.path.exists(f'{db_folder}room.json'):
        with open(f'{db_folder}room.json', 'w') as f:
            json.dump({}, f)
    room_data = {}
    with open(f'{db_folder}room.json', 'r') as f:
        room_data = json.load(f)
    room_data[code] = room_choice
    with open(f'{db_folder}room.json', 'w') as f:
        json.dump(room_data, f)

def useXL():
    if not os.path.exists(result_folder):
        os.makedirs(result_folder)
    with open(f'{db_folder}timetable.json', 'r') as file:
        timetable_data = json.load(file)

    for day_schedule in timetable_data:
        for slot in day_schedule:
            if slot["time"] == "02:00 - 02:55":
                slot["time"] = "14:00 - 14:55"
            elif slot["time"] == "03:00 - 03:55":
                slot["time"] = "15:00 - 15:55"
            elif slot["time"] == "04:00 - 04:55":
                slot["time"] = "16:00 - 16:55"
            elif slot["time"] == "05:00 - 05:55":
                slot["time"] = "17:00 - 17:55"
            
            if ',' in slot["items"][0]["room"]:
                room_choice = get_room_choice(slot["items"][0]["code"])
                if room_choice:
                    slot["items"][0]["room"] = room_choice
                else:
                    rooms = slot["items"][0]["room"].split(',')
                    print(f"Choose a room for the Course {slot["items"][0]["name"]} (by default we will keep all the rooms and ask you again):")
                    for i, room in enumerate(rooms):
                        print(f'{i+1}. {room.strip()}')
                    room_choice = input("Enter the room number: ")
                    append_room_choice(rooms[int(room_choice)-1].strip(), slot["items"][0]["code"])
                    slot["items"][0]["room"] = rooms[int(room_choice)-1]

    def time_key(time_str):
        start_time = time_str.split(" - ")[0]
        return datetime.strptime(start_time, "%H:%M")

    times = sorted({slot["time"] for day in timetable_data for slot in day}, key=time_key)

    for time in times:
        if time == "14:00 - 14:55":
            times[times.index(time)] = "02:00 - 02:55"
        elif time == "15:00 - 15:55":
            times[times.index(time)] = "03:00 - 03:55"
        elif time == "16:00 - 16:55":
            times[times.index(time)] = "04:00 - 04:55"
        elif time == "17:00 - 17:55":
            times[times.index(time)] = "05:00 - 05:55"

    for day_schedule in timetable_data:
        for slot in day_schedule:
            if slot["time"] == "14:00 - 14:55":
                slot["time"] = "02:00 - 02:55"
            elif slot["time"] == "15:00 - 15:55":
                slot["time"] = "03:00 - 03:55"
            elif slot["time"] == "16:00 - 16:55":
                slot["time"] = "04:00 - 04:55"
            elif slot["time"] == "17:00 - 17:55":
                slot["time"] = "05:00 - 05:55"

    day_order = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    days = sorted({slot["day"] for day in timetable_data for slot in day}, key=lambda d: day_order.index(d))

    max_content_length = 0
    schedule_dict = {}
    for day_schedule in timetable_data:
        for slot in day_schedule:
            key = (slot["day"], slot["time"])
            content = ''
            roomContent = ''
            if slot["items"][0]["room"] == "":
                roomContent = ''
            else:
                roomContent = f"Room - ({slot['items'][0]['room']})"
            if slot["items"][0]["name"] == "Free":
                content = ''
            else:
                content = f"{slot['items'][0]['name']}"
            content = content.upper()
            schedule_dict[key] = content+'\n'+roomContent
            max_content_length = max(max_content_length, len(content))
            max_content_length = max(max_content_length, len(roomContent))

    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable"

    col_width = max_content_length + 5
    ws.column_dimensions['A'].width = 10
    ws.row_dimensions[1].height = 30
    for col in range(2, len(times) + 2):
        ws.column_dimensions[chr(64 + col)].width = col_width

    header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    day_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    default_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    free_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    centered_alignment = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)

    ws.cell(row=1, column=1, value="")
    for col, time in enumerate(times, start=2):
        cell = ws.cell(row=1, column=col, value=time)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = centered_alignment
        cell.font = bold_font

    for row, day in enumerate(days, start=2):
        day_cell = ws.cell(row=row, column=1, value=day)
        day_cell.fill = day_fill
        day_cell.border = border
        day_cell.alignment = centered_alignment
        day_cell.font = bold_font
        for col, time in enumerate(times, start=2):
            text = schedule_dict.get((day, time), "")
            cell = ws.cell(row=row, column=col, value=text)
            if "\n" in text:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = centered_alignment
            if text == "\n":
                cell.fill = free_fill
            else:
                cell.fill = default_fill
            cell.border = border
            max_line_count = text.count("\n") + 1
            row_height = 15 * max_line_count
            ws.row_dimensions[row].height = row_height

    wb.save(f"{result_folder}timetable.xlsx")
    print("Timetable saved as timetable.xlsx in results folder")

    excel2img.export_img(f"{result_folder}timetable.xlsx", f"{result_folder}timetable.png")
    print("Timetable saved as timetable.png in results folder")